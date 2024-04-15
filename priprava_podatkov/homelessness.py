from openpyxl import load_workbook
import csv

# Absolutna pot do datoteke Excel
file_path = r'priprava_podatkov\originalni_podatki\2007-2023-PIT-Counts-by-State1.xlsx'
states_csv_path = r"podatki\states.csv"

# Seznam za shranjevanje samo kod držav
state_codes = []

# Preberi samo kode držav iz CSV datoteke
with open(states_csv_path, mode='r') as file:
    reader = csv.reader(file)
    next(reader)  # Preskoči glavo
    for row in reader:
        state, code = row
        state_codes.append(code.strip())

# Naloži Excel datoteko
wb = load_workbook(file_path, read_only=True)

# Pridobi seznam imen listov (sheetov)
sheet_names = wb.sheetnames
new_csv_path = r"podatki\homelessness.csv"
# Za vsak list v Excel datoteki
with open(new_csv_path, mode='w', newline='') as file:
    writer = csv.writer(file)

    # Dodaj naslov vrstice (header)
    writer.writerow(["Year", "State", "Number of CoCs", "Overall Homeless", "Sheltered ES Homeless", "Sheltered TH Homeless", "Sheltered Total Homeless", "Unsheltered Homeless"])

    # Za vsak list v Excel datoteki
    for sheet_name in wb.sheetnames:
        # Preverimo, ali ime lista začne z letnico
        if sheet_name[0].isdigit():
            ws = wb[sheet_name]

            # Za vsako vrstico v listu
            for row in ws.iter_rows(min_row=1, values_only=True):
                # Preverimo, ali je koda države v seznamu
                if row[0] in state_codes:
                    # Zapišemo prvih 6 stolpcev vrstice v novo CSV datoteko
                    writer.writerow([sheet_name, row[0], row[1], row[2], row[3], row[4], row[5], row[6]])
print("Ustvarjena je nova CSV datoteka.")
